import boto3
import botocore
# import json
import openpyxl
import csv

# import datetime3

ac_no = '405020847236'
role_arn = f'arn:aws:iam::{ac_no}:role/HPCORE-SUPPORT-ADMIN'
cli = boto3.session.Session(profile_name='default')
aws_con = cli.client('sts')
asume_role = (aws_con.assume_role(RoleArn=role_arn, RoleSessionName='pkvi'))
reduc = asume_role.get('Credentials')

aws_con2 = boto3.session.Session(aws_access_key_id=reduc['AccessKeyId'],
                                 aws_secret_access_key=reduc['SecretAccessKey'],
                                 aws_session_token=reduc['SessionToken'])

# Connect to CloudWatch
# cloudwatch = boto3.client("cloudwatch")

cloudwatch = aws_con2.client(service_name='cloudwatch', region_name='us-west-2')

# Define the alarm properties
metric_name = "disk_used_percent"
namespace = "CWAgent"
comparison_operator = "GreaterThanOrEqualToThreshold"
evaluation_periods = 2
threshold = 80.0
period = 300
statistic = "Average"

# Load Excel Workbook and select active sheet
file = input('Enter the path of the Excel WB: ')
wb = openpyxl.load_workbook(file)
# wb = openpyxl.load_workbook('D:\Users\pkvi\OneDrive - DXC Production\LB Logs\New folder\Book3.xlsx')
sheet = wb.active

# Create CSV file to store the output
with open('instance_alarms.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(["Instance ID", "Alarm Name", "Alarm Description", "Alarm Status"])

# Loop through each row in the Excel sheet
# maxrow=wb.max_row
# print(f"no.of rows {maxrow}")
for row in sheet.iter_rows(values_only=True):
    instance_id = row[0]
    if not instance_id:
        break

    device_type = row[1]
    fstype_value = row[2]

    Dimensions = [
        {
            "Name": "InstanceId",
            "Value": str(instance_id)
        },
        {
            "Name": "path",
            "Value": "/var"
        },
        {
            "Name": "device",
            "Value": str(device_type)
        },
        {
            "Name": "fstype",
            "Value": str(fstype_value)
        },
    ]

    # Create the alarm
    try:
        response = cloudwatch.put_metric_alarm(
            AlarmName='awsec2-' + str(instance_id) + '-/var-Disk-Pct-Used',
            ComparisonOperator=comparison_operator,
            EvaluationPeriods=evaluation_periods,
            MetricName=metric_name,
            Namespace=namespace,
            Period=period,
            Statistic=statistic,
            Threshold=threshold,
            AlarmDescription="Alarm for /var file system utilization",
            AlarmActions=[
                "arn:aws:sns:us-west-2:405020847236:DXC-TEST",
            ],
            Dimensions=Dimensions
        )
        Unit = "Percent"

        print(f"Alarm created for instance {instance_id}")
        with open('instance_alarms.csv', 'a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(
                [instance_id, f"awsec2-{instance_id}-/var-Disk-Pct-Used", "Alarm for /var file system utilization",
                 "Alarm created"])
    except botocore.exceptions.ClientError as e:
        print(f"Error creating alarm for instance {instance_id}: {e}")
